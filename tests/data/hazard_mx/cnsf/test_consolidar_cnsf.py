"""
Pruebas del consolidador CNSF (usa la muestra real 2024 de Incendio + fixtures
sintéticos para simular años viejos con deriva de estructura).

La muestra real vive en el data-root del repo (git-ignorado); si no está
descargada, las pruebas se saltan en vez de fallar.
"""

import shutil
import tempfile
from pathlib import Path

import consolidar_cnsf as C
import openpyxl
import pandas as pd
import pytest

_REPO = Path(__file__).resolve().parents[4]
BASE_2024 = (
    _REPO / "data" / "hazard_mx" / "datos_CNSF" / "crudos" / "incendio" / "2024 Incendio Bases.xlsx"
)
ROOT = Path(tempfile.gettempdir()) / "prueba_consol_cnsf"

pytestmark = pytest.mark.skipif(
    not BASE_2024.exists(),
    reason="muestra real CNSF ausente (data/hazard_mx/datos_CNSF/crudos/incendio)",
)


def _preparar():
    if ROOT.exists():
        shutil.rmtree(ROOT)
    d = ROOT / "incendio"
    d.mkdir(parents=True)
    shutil.copy(BASE_2024, d / "2024 Incendio Bases.xlsx")

    # 2007 simulado con deriva:
    wb = openpyxl.load_workbook(BASE_2024)
    wb["Emision"].title = "Emisión"  # nombre de hoja con acento (años viejos)
    ws = wb["Siniestros"]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v and "GIRO" in str(v):
            ws.cell(2, c).value = "GIRO DE LA UBICACIÓN"  # variante correcta vs typo 2024
    # columna en blanco con celdas vacías (spacer de SharePoint)
    nc = ws.max_column + 1
    ws.cell(2, nc).value = ""
    for s in list(wb.sheetnames):
        if s not in ("Indice", "Emisión", "Siniestros"):
            del wb[s]
    wb.save(d / "2007 Incendio Bases.xlsx")


def test_fusion_hojas_sin_colision():
    _preparar()
    out = ROOT / "out"
    C.consolidar(ROOT, out, categorias=["incendio"])
    # Debe existir UN solo emision.csv con AMBOS años (Emision + Emisión fusionadas)
    df = pd.read_csv(out / "incendio" / "emision.csv")
    anios = sorted(df["anio"].dropna().astype(int).unique().tolist())
    assert anios == [2007, 2024], anios
    print("OK fusión de hojas: emision.csv con años", anios)


def test_alias_por_ambito():
    _preparar()
    out = ROOT / "out_alias"
    aliases = {"incendio/siniestros": {"GIRO DE LA UBICACIÓN": "GIRO LA UBICACIÓN"}}
    C.consolidar(ROOT, out, categorias=["incendio"], aliases=aliases)
    sn = pd.read_csv(out / "incendio" / "siniestros.csv")
    # se fusionó: no queda la columna variante
    assert "GIRO DE LA UBICACIÓN" not in sn.columns
    assert "GIRO LA UBICACIÓN" in sn.columns
    # y la columna quedó poblada en ambos años
    pobl = sn.groupby("anio")["GIRO LA UBICACIÓN"].apply(lambda s: s.notna().sum()).to_dict()
    assert all(v > 0 for v in pobl.values()), pobl
    # el ámbito NO afectó a Emisión (su canónico propio se conserva)
    em = pd.read_csv(out / "incendio" / "emision.csv")
    assert "GIRO DE LA UBICACIÓN" in em.columns
    print("OK alias por ámbito: GIRO unificado en Siniestros, Emisión intacta")


def test_validacion_detecta_y_no_falsea():
    _preparar()
    out = ROOT / "out_val"
    import json

    C.consolidar(ROOT, out, categorias=["incendio"])
    rep = json.loads((out / "incendio" / "_reporte.json").read_text())
    sin = [h for h in rep["hojas"] if h["hoja"] == "Siniestros"][0]
    assert [h for h in rep["hojas"] if h["hoja"] in ("Emisión", "Emision")], rep["hojas"]
    # Siniestros: detecta el par GIRO como posible alias (variante histórica vs canónica)
    dups = [
        d
        for d in sin["validaciones"]["posibles_duplicados"]
        if "GIRO" in d["variante"] and "GIRO" in d["canonica"]
    ]
    assert dups, sin["validaciones"]["posibles_duplicados"]
    # columna en blanco fue descartada (no aparece como sin_encabezado en datos)
    assert sin["validaciones"]["columnas_sin_encabezado"] == [], sin["validaciones"]
    print("OK validación: sugiere alias GIRO y descarta columna en blanco")


if __name__ == "__main__":
    test_fusion_hojas_sin_colision()
    test_alias_por_ambito()
    test_validacion_detecta_y_no_falsea()
    print("\nTODAS LAS PRUEBAS DEL CONSOLIDADOR PASARON")
